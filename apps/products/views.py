# apps/products/views.py - COMPLETE FINAL VERSION

from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, TemplateView
from django.db.models import Q, Count, Min, Max, Prefetch, Avg
from django.core.paginator import Paginator
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.utils import timezone

from .models import Product, Category, Occasion, ProductImage, ProductVariant, CSVImportLog
from apps.products.services.csv_importer import CSVImporter


# ============================================
# PUBLIC PRODUCT VIEWS
# ============================================

class ProductListView(ListView):
    """Enhanced Product List View with filtering, sorting and pagination"""
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'
    paginate_by = 12

    def get_queryset(self):
        """Optimize queryset with proper prefetching"""
        queryset = Product.objects.filter(
            is_active=True,
            published=True,
            status='active'
        ).select_related('category').prefetch_related(
            Prefetch('images', queryset=ProductImage.objects.filter(is_active=True).order_by('sort_order', 'position'))
        )
        
        # Apply filters
        queryset = self.apply_filters(queryset)
        
        # Apply sorting
        queryset = self.apply_sorting(queryset)
        
        return queryset

    def apply_filters(self, queryset):
        """Apply various filters to the queryset"""
        
        # Quick filters
        filter_type = self.request.GET.get('filter')
        if filter_type == 'featured':
            queryset = queryset.filter(is_featured=True)
        elif filter_type == 'bestseller':
            queryset = queryset.filter(is_bestseller=True)
        elif filter_type == 'under_500':
            queryset = queryset.filter(base_price__lt=500)
        elif filter_type == 'under_1000':
            queryset = queryset.filter(base_price__lt=1000)
        elif filter_type == 'under_2000':
            queryset = queryset.filter(base_price__lt=2000)

        # Price range filters
        price_ranges = self.request.GET.getlist('price_range')
        if price_ranges:
            price_q = Q()
            for price_range in price_ranges:
                if price_range == '0-500':
                    price_q |= Q(base_price__lt=500)
                elif price_range == '500-1000':
                    price_q |= Q(base_price__gte=500, base_price__lt=1000)
                elif price_range == '1000-2000':
                    price_q |= Q(base_price__gte=1000, base_price__lt=2000)
                elif price_range == '2000+':
                    price_q |= Q(base_price__gte=2000)
            queryset = queryset.filter(price_q)

        # Category filters
        categories = self.request.GET.getlist('categories')
        if categories:
            queryset = queryset.filter(category__slug__in=categories)

        # Occasion filters
        occasions = self.request.GET.getlist('occasions')
        if occasions:
            queryset = queryset.filter(occasions__slug__in=occasions)

        # Vendor filter (from CSV data)
        vendors = self.request.GET.getlist('vendors')
        if vendors:
            queryset = queryset.filter(vendor__in=vendors)

        # Brand filter (from CSV data)
        brands = self.request.GET.getlist('brands')
        if brands:
            queryset = queryset.filter(brand__in=brands)

        # Stock filter
        stock_filter = self.request.GET.get('stock')
        if stock_filter == 'in_stock':
            queryset = queryset.filter(stock_quantity__gt=0)
        elif stock_filter == 'out_of_stock':
            queryset = queryset.filter(stock_quantity=0)

        return queryset

    def apply_sorting(self, queryset):
        """Apply sorting to the queryset"""
        sort_by = self.request.GET.get('sort', 'popularity')
        
        if sort_by == 'price_low':
            return queryset.order_by('base_price')
        elif sort_by == 'price_high':
            return queryset.order_by('-base_price')
        elif sort_by == 'newest':
            return queryset.order_by('-created_at')
        elif sort_by == 'name_asc':
            return queryset.order_by('name')
        elif sort_by == 'name_desc':
            return queryset.order_by('-name')
        else:  # popularity (default)
            return queryset.order_by('-is_featured', '-is_bestseller', '-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add page title
        context['page_title'] = 'All Products'

        # Add search query context
        search_query = self.request.GET.get('q', '').strip()
        if search_query:
            context['search_query'] = search_query
            context['page_title'] = f'Search Results for "{search_query}"'

        # Add filter context
        try:
            context['available_categories'] = Category.objects.filter(
                products__is_active=True,
                products__published=True
            ).distinct()
        except:
            context['available_categories'] = Category.objects.none()

        try:
            context['available_occasions'] = Occasion.objects.filter(
                products__is_active=True,
                products__published=True
            ).distinct()
        except:
            context['available_occasions'] = Occasion.objects.none()
        
        # Get available vendors and brands from CSV data
        try:
            context['available_vendors'] = Product.objects.filter(
                is_active=True, 
                published=True,
                vendor__isnull=False
            ).exclude(vendor='').values_list('vendor', flat=True).distinct()
        except:
            context['available_vendors'] = []

        try:
            context['available_brands'] = Product.objects.filter(
                is_active=True,
                published=True,
                brand__isnull=False
            ).exclude(brand='').values_list('brand', flat=True).distinct()
        except:
            context['available_brands'] = []

        # Add product count
        context['total_products'] = self.get_queryset().count()

        # Current filters (for active state in UI)
        context['active_filters'] = {
            'price_ranges': self.request.GET.getlist('price_range'),
            'categories': self.request.GET.getlist('categories'),
            'occasions': self.request.GET.getlist('occasions'),
            'vendors': self.request.GET.getlist('vendors'),
            'brands': self.request.GET.getlist('brands'),
            'sort': self.request.GET.get('sort', 'popularity'),
        }

        return context


class ProductDetailView(DetailView):
    """Enhanced Product Detail View"""
    model = Product
    template_name = 'products/product_detail.html'
    context_object_name = 'product'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'

    def get_queryset(self):
        return Product.objects.filter(
            is_active=True,
            published=True
        ).select_related('category').prefetch_related(
            'images',
            'variants',
            'occasions'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Add all product images in order
        context['product_images'] = self.object.all_images

        # Add product variants
        context['product_variants'] = self.object.variants.filter(is_active=True).order_by('sort_order')

        # Calculate discount percentage
        context['discount_percentage'] = self.object.discount_percentage

        # FIX: Get related products from same category (increased to 6)
        from django.db import models
        related_products = Product.objects.filter(
            category=self.object.category,
            is_active=True,
            published=True,
            stock_quantity__gt=0
        ).exclude(
            pk=self.object.pk
        ).select_related('category').prefetch_related('images')[:6]

        # DEBUG logging
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Found {related_products.count()} related products for {self.object.name}")

        context['related_products'] = related_products

        # BONUS: recommended products (bestsellers or featured from same category)
        context['recommended_products'] = Product.objects.filter(
            category=self.object.category,
            is_active=True,
            published=True,
            stock_quantity__gt=0
        ).filter(
            models.Q(is_bestseller=True) | models.Q(is_featured=True)
        ).exclude(pk=self.object.pk)[:4]

        # Check stock status
        context['in_stock'] = self.object.is_in_stock
        context['low_stock'] = 0 < self.object.stock_quantity <= 5

        # Add shipping info
        context['free_shipping'] = self.object.current_price >= 500

        # Add recommended categories (featured categories with products)
        context['recommended_categories'] = Category.objects.filter(
            is_featured=True,
            is_active=True,
            products__is_active=True
        ).distinct()[:6]

        # Add bestseller products
        context['bestseller_products'] = Product.objects.filter(
            is_bestseller=True,
            is_active=True,
            published=True
        ).select_related('category').prefetch_related('images')[:8]

        # If no bestsellers, show featured products instead
        if not context['bestseller_products'].exists():
            context['bestseller_products'] = Product.objects.filter(
                is_featured=True,
                is_active=True,
                published=True
            ).select_related('category').prefetch_related('images')[:8]

        # Add available add-ons
        context['available_addons'] = self.object.get_available_addons()

        # Add product reviews
        try:
            from apps.reviews.models import Review
            reviews = Review.objects.filter(
                product=self.object,
                is_approved=True,
                is_active=True
            ).select_related('user')[:10]

            context['product_reviews'] = reviews
            context['reviews_count'] = reviews.count()

            if reviews.exists():
                avg = sum([r.rating for r in reviews]) / reviews.count()
                context['average_rating'] = round(avg, 1)
        except:
            context['product_reviews'] = []
            context['reviews_count'] = 0
            context['average_rating'] = 0

        return context


class CategoryDetailView(ListView):
    """Category-specific product listing"""
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'
    paginate_by = 12

    def get_queryset(self):
        self.category = get_object_or_404(Category, slug=self.kwargs['slug'], is_active=True)
        queryset = Product.objects.filter(
            category=self.category,
            is_active=True,
            published=True
        ).select_related('category').prefetch_related('images')
        
        # Apply the same filtering and sorting as ProductListView
        view = ProductListView()
        view.request = self.request
        view.kwargs = self.kwargs
        queryset = view.apply_filters(queryset)
        queryset = view.apply_sorting(queryset)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category'] = self.category
        context['page_title'] = self.category.name

        # Add the same context as ProductListView
        try:
            context['available_categories'] = Category.objects.filter(
                products__is_active=True,
                products__published=True
            ).distinct()
        except:
            context['available_categories'] = Category.objects.none()

        try:
            context['available_occasions'] = Occasion.objects.filter(
                products__is_active=True,
                products__published=True
            ).distinct()
        except:
            context['available_occasions'] = Occasion.objects.none()

        context['total_products'] = self.get_queryset().count()

        return context


class OccasionDetailView(ListView):
    """Occasion-specific product listing"""
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'
    paginate_by = 12

    def get_queryset(self):
        slug = self.kwargs['slug']

        # Try to get exact Occasion match
        try:
            self.occasion = Occasion.objects.get(slug=slug, is_active=True)
            queryset = Product.objects.filter(
                occasions=self.occasion,
                is_active=True,
                published=True
            )
        except Occasion.DoesNotExist:
            # Fallback: Search by slug keywords in product name/description
            self.occasion = None
            search_terms = slug.replace('-', ' ')
            queryset = Product.objects.filter(
                Q(name__icontains=search_terms) |
                Q(category__name__icontains=search_terms) |
                Q(description__icontains=search_terms),
                is_active=True,
                published=True
            ).distinct()

        queryset = queryset.select_related('category').prefetch_related('images')

        # Apply the same filtering and sorting as ProductListView
        view = ProductListView()
        view.request = self.request
        view.kwargs = self.kwargs
        queryset = view.apply_filters(queryset)
        queryset = view.apply_sorting(queryset)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.occasion:
            context['occasion'] = self.occasion
            context['page_title'] = self.occasion.name
        else:
            context['page_title'] = self.kwargs['slug'].replace('-', ' ').title()

        # Add the same context as ProductListView
        try:
            context['available_categories'] = Category.objects.filter(
                products__is_active=True,
                products__published=True
            ).distinct()
        except:
            context['available_categories'] = Category.objects.none()

        try:
            context['available_occasions'] = Occasion.objects.filter(
                products__is_active=True,
                products__published=True
            ).distinct()
        except:
            context['available_occasions'] = Occasion.objects.none()

        context['total_products'] = self.get_queryset().count()

        return context


def product_search(request):
    """Enhanced search with CSV product data"""
    query = request.GET.get('q', '').strip()
    products = Product.objects.filter(
        is_active=True,
        published=True
    ).select_related('category').prefetch_related('images')

    if query:
        products = products.filter(
            Q(name__icontains=query) |
            Q(description__icontains=query) |
            Q(body_html__icontains=query) |
            Q(sku__icontains=query) |
            Q(category__name__icontains=query) |
            Q(brand__icontains=query) |
            Q(vendor__icontains=query) |
            Q(tags__icontains=query)
        ).distinct()

    # Apply the same filtering and sorting as ProductListView
    view = ProductListView()
    view.request = request
    view.kwargs = {}
    products = view.apply_filters(products)
    products = view.apply_sorting(products)

    # Pagination
    paginator = Paginator(products, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'products': page_obj,
        'search_query': query,
        'page_title': f'Search Results for "{query}"' if query else 'Search',
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
        'total_products': products.count(),
        'available_categories': Category.objects.filter(
            products__is_active=True,
            products__published=True
        ).distinct(),
        'available_occasions': Occasion.objects.filter(
            products__is_active=True,
            products__published=True
        ).distinct(),
    }

    return render(request, 'products/product_list.html', context)


# ============================================
# CSV IMPORT VIEWS (Admin Only)
# ============================================

@staff_member_required
def import_csv_view(request):
    """CSV/Excel import interface for admin users"""
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        file_type = request.POST.get('file_type', 'auto')
        
        if not csv_file:
            messages.error(request, 'Please select a file to upload.')
            return redirect(reverse('admin:import_csv'))
        
        # Check file extension
        allowed_extensions = ['.csv', '.xlsx', '.xls']
        file_ext = '.' + csv_file.name.split('.')[-1].lower()
        if file_ext not in allowed_extensions:
            messages.error(request, 'Please upload a valid CSV or Excel file (.csv, .xlsx, .xls)')
            return redirect(reverse('admin:import_csv'))
        
        # Check file size (10MB limit)
        if csv_file.size > 10 * 1024 * 1024:
            messages.error(request, 'File too large. Maximum size is 10MB.')
            return redirect(reverse('admin:import_csv'))
        
        # Convert Excel to CSV if needed
        if file_ext in ['.xlsx', '.xls']:
            try:
                import openpyxl
                import io
                import csv as csv_module
                
                # Read Excel file
                wb = openpyxl.load_workbook(csv_file)
                ws = wb.active
                
                # Convert to CSV in memory
                output = io.StringIO()
                csv_writer = csv_module.writer(output)
                for row in ws.iter_rows(values_only=True):
                    csv_writer.writerow(row)
                
                # Create a new file-like object with CSV data
                output.seek(0)
                csv_file = io.BytesIO(output.getvalue().encode('utf-8'))
                csv_file.name = csv_file.name.rsplit('.', 1)[0] + '.csv'
            except ImportError:
                messages.error(request, 'Excel support not installed. Please install openpyxl: pip install openpyxl')
                return redirect(reverse('admin:import_csv'))
            except Exception as e:
                messages.error(request, f'Error reading Excel file: {str(e)}')
                return redirect(reverse('admin:import_csv'))
        
        # Process the import
        try:
            importer = CSVImporter(
                user=request.user,
                file_obj=csv_file,
                file_type=file_type
            )
            result = importer.import_csv()
            
            if result['success']:
                messages.success(
                    request,
                    f"✓ Import completed! Created: {result['created']}, Updated: {result['updated']}"
                )
                if result['errors']:
                    messages.warning(
                        request,
                        f"⚠ {len(result['errors'])} errors occurred. Check import log for details."
                    )
            else:
                messages.error(request, f"✗ Import failed: {result['error']}")
        except Exception as e:
            messages.error(request, f"Unexpected error: {str(e)}")
        
        return redirect(reverse('admin:import_csv'))
    
    # GET request - show import form
    recent_imports = CSVImportLog.objects.all()[:10]
    context = {
        'title': 'Bulk Product Upload',
        'recent_imports': recent_imports,
        'has_permission': True,
        'site_header': 'GiftTree Administration',
        'site_title': 'Admin',
        'total_products': Product.objects.count(),
        'active_products': Product.objects.filter(published=True, is_active=True).count(),
    }
    return render(request, 'admin/import_csv.html', context)


@staff_member_required
@require_http_methods(["POST"])
def reset_database_view(request):
    """Reset product database (admin only)"""
    confirmation = request.POST.get('confirmation', '')
    
    if confirmation != 'RESET':
        return JsonResponse({
            'success': False,
            'error': 'Please type RESET to confirm database reset'
        }, status=400)
    
    try:
        # Delete all products (cascades to images and variants)
        product_count = Product.objects.count()
        Product.objects.all().delete()
        
        # Optionally delete categories
        if request.POST.get('delete_categories') == 'true':
            Category.objects.all().delete()
        
        # Optionally delete import logs
        if request.POST.get('delete_logs') == 'true':
            CSVImportLog.objects.all().delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully deleted {product_count} products'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@staff_member_required
def download_template(request):
    """Download sample CSV/Excel template for bulk upload"""
    import csv
    from django.http import HttpResponse
    
    format_type = request.GET.get('format', 'gifttree')  # gifttree or shopify
    file_type = request.GET.get('type', 'csv')  # csv or excel
    
    # Define template headers and sample data
    if format_type == 'gifttree':
        headers = [
            'SKU', 'Name', 'Description', 'Category', 'Base Price', 'Sale Price', 
            'Discount Price', 'MRP', 'Stock Quantity', 'Weight', 'Size', 'Color',
            'Brand', 'Vendor', 'Tags', 'Is Featured', 'Is Bestseller', 'Is Active',
            'Meta Title', 'Meta Description', 'Image1', 'Image2', 'Image3'
        ]
        sample_data = [
            [
                'GIFT-001', 
                'Chocolate Gift Box', 
                'Delicious assorted chocolates in a premium gift box', 
                'Chocolates', 
                '999.00', 
                '899.00', 
                '799.00', 
                '1099.00', 
                '50', 
                '0.5', 
                'Medium', 
                'Brown', 
                'Sweet Delights', 
                'Gift Vendor', 
                'chocolate,gift,premium', 
                'TRUE', 
                'TRUE', 
                'TRUE', 
                'Premium Chocolate Gift Box - Best Gift for Any Occasion',
                'Buy premium chocolate gift box with assorted chocolates. Perfect gift for birthdays, anniversaries.',
                'https://example.com/image1.jpg',
                'https://example.com/image2.jpg',
                'https://example.com/image3.jpg'
            ],
            [
                'GIFT-002', 
                'Red Roses Bouquet', 
                'Beautiful bouquet of 12 fresh red roses', 
                'Flowers', 
                '599.00', 
                '549.00', 
                '499.00', 
                '699.00', 
                '100', 
                '0.3', 
                'Standard', 
                'Red', 
                'Fresh Flowers Co', 
                'Flower Vendor', 
                'roses,flowers,romantic,gift', 
                'TRUE', 
                'FALSE', 
                'TRUE', 
                '12 Red Roses Bouquet - Express Your Love',
                'Fresh red roses bouquet, perfect for expressing love and affection. Same day delivery available.',
                'https://example.com/roses1.jpg',
                'https://example.com/roses2.jpg',
                ''
            ]
        ]
    else:  # Shopify format
        headers = [
            'Handle', 'Title', 'Body (HTML)', 'Vendor', 'Product Category', 'Type', 
            'Tags', 'Published', 'Option1 Name', 'Option1 Value', 'Variant SKU', 
            'Variant Price', 'Variant Compare At Price', 'Variant Inventory Qty', 
            'Variant Weight', 'Variant Weight Unit', 'Variant Taxable', 
            'Image Src', 'Image Position', 'SEO Title', 'SEO Description', 'Status'
        ]
        sample_data = [
            [
                'chocolate-gift-box',
                'Chocolate Gift Box',
                '<p>Delicious assorted chocolates in a premium gift box</p>',
                'Sweet Delights',
                'Gifts',
                'Chocolates',
                'chocolate, gift, premium',
                'TRUE',
                'Size',
                'Medium',
                'CHO-001-M',
                '899.00',
                '1099.00',
                '50',
                '0.5',
                'kg',
                'TRUE',
                'https://example.com/image1.jpg',
                '1',
                'Premium Chocolate Gift Box',
                'Buy premium chocolate gift box with assorted chocolates',
                'active'
            ],
            [
                'red-roses-bouquet',
                'Red Roses Bouquet',
                '<p>Beautiful bouquet of 12 fresh red roses</p>',
                'Fresh Flowers Co',
                'Gifts',
                'Flowers',
                'roses, flowers, romantic, gift',
                'TRUE',
                'Quantity',
                '12 Roses',
                'ROS-002-12',
                '549.00',
                '699.00',
                '100',
                '0.3',
                'kg',
                'TRUE',
                'https://example.com/roses1.jpg',
                '1',
                '12 Red Roses Bouquet',
                'Fresh red roses bouquet, perfect for expressing love',
                'active'
            ]
        ]
    
    # Generate CSV or Excel
    if file_type == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # Add headers with styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Add sample data
            for row_num, row_data in enumerate(sample_data, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{format_type}_products_template.xlsx"'
            wb.save(response)
            return response
            
        except ImportError:
            messages.error(request, 'Excel support not installed. Downloading as CSV instead.')
            file_type = 'csv'
    
    # Generate CSV (default or fallback)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{format_type}_products_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in sample_data:
        writer.writerow(row)
    
    return response


# ============================================
# API ENDPOINTS (for Ajax calls)
# ============================================

@require_http_methods(["GET"])
def product_list_api(request):
    """API endpoint for product listing - used by Ajax"""
    products = Product.objects.filter(
        published=True,
        is_active=True,
        status='active'
    ).select_related('category').prefetch_related('images')[:100]
    
    data = []
    for product in products:
        primary_image = product.primary_image
        
        data.append({
            'id': product.id,
            'sku': product.sku,
            'name': product.name,
            'slug': product.slug,
            'price': float(product.current_price),
            'mrp': float(product.mrp) if product.mrp else None,
            'category': product.category.name if product.category else None,
            'image': primary_image.get_image_url if primary_image else None,
            'in_stock': product.is_in_stock,
            'vendor': product.vendor,
            'brand': product.brand,
        })
    
    return JsonResponse({'products': data, 'count': len(data)})


@require_http_methods(["GET"])
def product_detail_api(request, slug):
    """API endpoint for single product - used by Ajax quick view"""
    try:
        product = Product.objects.select_related('category').prefetch_related(
            'images', 'variants', 'occasions'
        ).get(slug=slug, published=True, is_active=True)
        
        data = {
            'id': product.id,
            'sku': product.sku,
            'name': product.name,
            'slug': product.slug,
            'description': product.body_html or product.description,
            'price': float(product.current_price),
            'mrp': float(product.mrp) if product.mrp else None,
            'discount_percentage': product.discount_percentage,
            'category': product.category.name if product.category else None,
            'vendor': product.vendor,
            'brand': product.brand,
            'tags': product.tags.split(',') if product.tags else [],
            'images': [
                {
                    'url': img.get_image_url,
                    'position': img.position or img.sort_order,
                    'alt': img.alt_text,
                    'is_primary': img.is_primary
                } for img in product.all_images
            ],
            'variants': [
                {
                    'id': var.id,
                    'name': var.display_name,
                    'sku': var.full_sku,
                    'price': float(var.final_price),
                    'in_stock': var.stock_quantity > 0,
                } for var in product.variants.filter(is_active=True)
            ],
            'in_stock': product.is_in_stock,
            'stock_quantity': product.stock_quantity,
        }
        
        return JsonResponse({'success': True, 'product': data})
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found'}, status=404)


@require_http_methods(["GET"])
def product_autocomplete_api(request):
    """Autocomplete API for search suggestions"""
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({'suggestions': []})
    
    products = Product.objects.filter(
        Q(name__icontains=query) | Q(sku__icontains=query),
        is_active=True,
        published=True
    ).select_related('category').prefetch_related('images')[:10]
    
    suggestions = []
    for product in products:
        primary_image = product.primary_image
        suggestions.append({
            'id': product.id,
            'name': product.name,
            'slug': product.slug,
            'price': float(product.current_price),
            'image': primary_image.get_image_url if primary_image else None,
            'category': product.category.name if product.category else None,
        })
    
    return JsonResponse({'suggestions': suggestions})


# ============================================
# NEW MENU-BASED VIEWS
# ============================================

class MenuCategoryDetailView(ListView):
    """Menu Category-specific product listing"""
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'
    paginate_by = 12

    def get_queryset(self):
        from .models import MenuCategory
        self.menu_category = get_object_or_404(MenuCategory, slug=self.kwargs['slug'], is_active=True)
        
        # Try to find corresponding old category
        try:
            old_category = Category.objects.get(slug=self.menu_category.slug, is_active=True)
            queryset = Product.objects.filter(
                category=old_category,
                is_active=True,
                published=True
            ).select_related('category').prefetch_related('images')
        except Category.DoesNotExist:
            queryset = Product.objects.none()
        
        # Apply filtering and sorting
        view = ProductListView()
        view.request = self.request
        view.kwargs = self.kwargs
        queryset = view.apply_filters(queryset)
        queryset = view.apply_sorting(queryset)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['menu_category'] = self.menu_category
        context['page_title'] = self.menu_category.name
        return context


class ProductTypeDetailView(ListView):
    """Product Type-specific product listing with fallback search"""
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'
    paginate_by = 12

    def get_queryset(self):
        from .models import ProductType
        slug = self.kwargs['slug']

        # Try to get exact ProductType match
        try:
            self.product_type = ProductType.objects.get(slug=slug, is_active=True)
            queryset = Product.objects.filter(
                product_types=self.product_type,
                is_active=True,
                published=True
            )
        except ProductType.DoesNotExist:
            # Fallback: Search by slug keywords in product name/category
            self.product_type = None
            search_terms = slug.replace('-', ' ')
            queryset = Product.objects.filter(
                Q(name__icontains=search_terms) |
                Q(category__name__icontains=search_terms) |
                Q(description__icontains=search_terms),
                is_active=True,
                published=True
            ).distinct()

        queryset = queryset.select_related('category').prefetch_related('images')

        # Apply filtering and sorting
        view = ProductListView()
        view.request = self.request
        view.kwargs = self.kwargs
        queryset = view.apply_filters(queryset)
        queryset = view.apply_sorting(queryset)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.product_type:
            context['product_type'] = self.product_type
            context['page_title'] = self.product_type.name
        else:
            context['page_title'] = self.kwargs['slug'].replace('-', ' ').title()
        return context


class CollectionDetailView(ListView):
    """Collection-specific product listing with fallback search"""
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'
    paginate_by = 12

    def get_queryset(self):
        from .models import Collection
        slug = self.kwargs['slug']

        # Try to get exact Collection match
        try:
            self.collection = Collection.objects.get(slug=slug, is_active=True)
            queryset = Product.objects.filter(
                collections=self.collection,
                is_active=True,
                published=True
            )
        except Collection.DoesNotExist:
            # Fallback: Search by slug keywords
            self.collection = None
            search_terms = slug.replace('-', ' ')
            queryset = Product.objects.filter(
                Q(name__icontains=search_terms) |
                Q(category__name__icontains=search_terms) |
                Q(description__icontains=search_terms),
                is_active=True,
                published=True
            ).distinct()

        queryset = queryset.select_related('category').prefetch_related('images')

        # Apply filtering and sorting
        view = ProductListView()
        view.request = self.request
        view.kwargs = self.kwargs
        queryset = view.apply_filters(queryset)
        queryset = view.apply_sorting(queryset)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.collection:
            context['collection'] = self.collection
            context['page_title'] = self.collection.name
        else:
            context['page_title'] = self.kwargs['slug'].replace('-', ' ').title()
        return context


class RecipientDetailView(ListView):
    """Recipient-specific product listing with fallback search"""
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'
    paginate_by = 12

    def get_queryset(self):
        from .models import Recipient
        slug = self.kwargs['slug']

        # Try to get exact Recipient match
        try:
            self.recipient = Recipient.objects.get(slug=slug, is_active=True)
            queryset = Product.objects.filter(
                recipients=self.recipient,
                is_active=True,
                published=True
            )
        except Recipient.DoesNotExist:
            # Fallback: Search by slug keywords
            self.recipient = None
            search_terms = slug.replace('-', ' ').replace('for-', '')
            queryset = Product.objects.filter(
                Q(name__icontains=search_terms) |
                Q(category__name__icontains=search_terms) |
                Q(description__icontains=search_terms),
                is_active=True,
                published=True
            ).distinct()

        queryset = queryset.select_related('category').prefetch_related('images')

        # Apply filtering and sorting
        view = ProductListView()
        view.request = self.request
        view.kwargs = self.kwargs
        queryset = view.apply_filters(queryset)
        queryset = view.apply_sorting(queryset)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.recipient:
            context['recipient'] = self.recipient
            context['page_title'] = f'Gifts for {self.recipient.name}'
        else:
            name = self.kwargs['slug'].replace('-', ' ').replace('for ', '').title()
            context['page_title'] = f'Gifts for {name}'
        return context


class LocationDetailView(ListView):
    """Location-specific product listing with fallback to all products"""
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'
    paginate_by = 12

    def get_queryset(self):
        from .models import DeliveryLocation
        slug = self.kwargs['slug']

        # Try to get exact DeliveryLocation match
        try:
            self.location = DeliveryLocation.objects.get(slug=slug, is_active=True)
            queryset = Product.objects.filter(
                delivery_locations=self.location,
                is_active=True,
                published=True
            )
        except DeliveryLocation.DoesNotExist:
            # Fallback: Show all products (most products are deliverable everywhere)
            self.location = None
            queryset = Product.objects.filter(
                is_active=True,
                published=True
            )

        queryset = queryset.select_related('category').prefetch_related('images')

        # Apply filtering and sorting
        view = ProductListView()
        view.request = self.request
        view.kwargs = self.kwargs
        queryset = view.apply_filters(queryset)
        queryset = view.apply_sorting(queryset)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.location:
            context['location'] = self.location
            context['page_title'] = f'Delivery to {self.location.name}'
        else:
            location_name = self.kwargs['slug'].replace('-', ' ').title()
            context['page_title'] = f'Delivery to {location_name}'
        return context


class CakeCategoryView(TemplateView):
    """Cakes category page with subcategories"""
    template_name = 'products/cakes_category.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        try:
            cakes_category = Category.objects.get(slug='cakes', is_active=True)
        except Category.DoesNotExist:
            cakes_category = None

        context['cakes_category'] = cakes_category

        if cakes_category:
            context['cake_subcategories'] = Category.objects.filter(
                parent=cakes_category,
                is_active=True
            ).order_by('sort_order')

        context['featured_cakes'] = Product.objects.filter(
            is_featured=True,
            is_active=True
        ).select_related('category').prefetch_related('images')[:8]

        return context


@require_http_methods(["GET"])
def quick_view(request, product_id):
    """Quick view API for product details"""
    try:
        product = Product.objects.select_related('category').prefetch_related(
            'images',
            'variants',
            'occasions'
        ).get(id=product_id, is_active=True, published=True)

        # Get product images
        images = []
        for img in product.all_images:
            images.append({
                'url': img.get_image_url,
                'alt': img.alt_text or product.name,
                'is_primary': img.is_primary
            })

        # Get product variants
        variants = []
        for variant in product.variants.filter(is_active=True):
            variants.append({
                'id': variant.id,
                'name': variant.display_name,
                'price': float(variant.final_price),
                'stock': variant.stock_quantity,
                'sku': variant.full_sku
            })

        # Build response data
        data = {
            'success': True,
            'product': {
                'id': product.id,
                'name': product.name,
                'slug': product.slug,
                'description': product.description,
                'price': float(product.current_price),
                'discount_price': float(product.discount_price) if product.discount_price else None,
                'discount_percentage': product.discount_percentage,
                'stock_quantity': product.stock_quantity,
                'is_in_stock': product.is_in_stock,
                'category': {
                    'name': product.category.name,
                    'slug': product.category.slug
                } if product.category else None,
                'images': images,
                'variants': variants,
                'customization': {
                    'allow_customization': product.allow_customization,
                    'allow_name_customization': product.allow_name_customization,
                    'allow_message_customization': product.allow_message_customization,
                    'allow_flavor_selection': product.allow_flavor_selection,
                    'customization_note': product.customization_note
                }
            }
        }

        return JsonResponse(data)

    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Product not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_http_methods(["GET"])
def check_pincode_availability(request):
    """
    API endpoint to check if a product is available for a given pincode
    GET params: pincode, product_id, variant_id (optional)
    """
    pincode = request.GET.get('pincode', '').strip()
    product_id = request.GET.get('product_id')
    variant_id = request.GET.get('variant_id')

    if not pincode or not product_id:
        return JsonResponse({
            'success': False,
            'error': 'Pincode and product_id are required'
        }, status=400)

    # Validate pincode format (6 digits)
    if not pincode.isdigit() or len(pincode) != 6:
        return JsonResponse({
            'success': False,
            'available': False,
            'message': 'Please enter a valid 6-digit pincode'
        })

    try:
        product = Product.objects.get(id=product_id, is_active=True, published=True)
        variant = None

        if variant_id:
            try:
                variant = ProductVariant.objects.get(id=variant_id, product=product, is_active=True)
            except ProductVariant.DoesNotExist:
                pass

        # Check availability
        is_available, seller_location, available_quantity, delivery_days = product.check_availability_by_pincode(
            pincode, variant
        )

        if is_available:
            return JsonResponse({
                'success': True,
                'available': True,
                'message': f'Available! Delivery in {delivery_days} days',
                'delivery_days': delivery_days,
                'seller_location': {
                    'name': seller_location.name,
                    'city': seller_location.city,
                    'state': seller_location.state
                } if seller_location else None,
                'pincode_info': {
                    'pincode': pincode,
                    'is_serviceable': True
                }
            })
        else:
            return JsonResponse({
                'success': True,
                'available': False,
                'message': 'Sorry, this product is not available for your pincode',
                'pincode_info': {
                    'pincode': pincode,
                    'is_serviceable': False
                }
            })

    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Product not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_http_methods(["GET"])
def validate_pincode(request):
    """
    API endpoint to validate and get pincode details
    GET params: pincode
    """
    from apps.users.models import Pincode

    pincode = request.GET.get('pincode', '').strip()

    if not pincode:
        return JsonResponse({
            'success': False,
            'error': 'Pincode is required'
        }, status=400)

    # Validate pincode format
    if not pincode.isdigit() or len(pincode) != 6:
        return JsonResponse({
            'success': False,
            'valid': False,
            'message': 'Please enter a valid 6-digit pincode'
        })

    try:
        pincode_obj = Pincode.objects.get(pincode=pincode)

        return JsonResponse({
            'success': True,
            'valid': True,
            'serviceable': pincode_obj.is_serviceable and pincode_obj.is_active,
            'pincode_data': {
                'pincode': pincode_obj.pincode,
                'area': pincode_obj.area,
                'city': pincode_obj.city,
                'district': pincode_obj.district,
                'state': pincode_obj.state,
                'delivery_days': pincode_obj.delivery_days
            },
            'message': f'{pincode_obj.city}, {pincode_obj.state}' if pincode_obj.is_serviceable else 'Pincode not serviceable'
        })

    except Pincode.DoesNotExist:
        return JsonResponse({
            'success': True,
            'valid': False,
            'serviceable': False,
            'message': 'Pincode not found in our database'
        })